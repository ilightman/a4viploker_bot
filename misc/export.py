import io
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook

from db_api.db_services import Record


async def export_records_to_excel(records: list[Record]) -> io.BytesIO:
    """Экспорт всех записей в Excel файл в BytesIO"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Список кодов'
    titles = ['Номер', 'Код', 'Активирован', 'Инфо о клиенте', 'Дата отправки', 'Трек номер', 'Дата вручения',
              'Дата активации']
    ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], ws['F1'], ws['G1'], ws['H1'] = titles
    for row_number, record in enumerate(records, 2):
        ws[f'A{row_number}'] = record.code_number
        ws[f'B{row_number}'] = record.key
        ws[f'C{row_number}'] = 'Да' if record.is_activated else ''
        ws[f'D{row_number}'] = f'''{record.name}
        {record.phone}
        {record.address}
        {record.telegram_id}
        {record.username}
        '''
        ws[f'E{row_number}'] = record.date_of_shipping if record.date_of_shipping != '-' else ''
        ws[f'F{row_number}'] = record.track_number if record.track_number != '-' else ''
        ws[f'G{row_number}'] = record.date_of_receiving if record.date_of_receiving != '-' else ''
        ws[f'H{row_number}'] = record.activation_date if record.activation_date != '-' else ''
    xlsx_file = io.BytesIO()
    now_time = datetime.now(tz=ZoneInfo('Europe/Moscow')).strftime("%d.%m.%y, %H:%M:%S")
    xlsx_file.name = f'{now_time}.xlsx'
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file
